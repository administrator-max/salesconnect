#!/usr/bin/env python3
"""Generate iqdash/data/quotaLedger.json from the IQ Dash master workbook.

WHY THIS EXISTS
---------------
quotaLedger.json drives every headline quota KPI on the dashboard (obtained,
utilization, available) via iq_apply_ledger() in iqdash_data.php. It was
hand-built on 2026-07-01 from master copy (6) and had no generator, so it
silently froze while the master moved on: by 2026-07-27 the dashboard read
33,730 / 18,346 against a master at 34,240 / 22,547.

MODEL
-----
The master's `Status Submisson` sheet is a block per company. Row 2 carries
product names, row 3 the HS codes; each block's rows are labelled in column C:

    Submit #N            applied-for volume        (ignored - not granted)
    Obtained #N          granted volume            -> obtained
    Revision #N          +/- moves between products -> obtained (signed)
    Utilization (MT)     allocated to customers    -> util
    Available (MT)       obtained - util           (derived, ignored)

`_meta.view` is "effective (obtained incl. revisions)", so Revision rows are
summed into obtained. A company block starts where column A holds a number and
column B a name; the code is the leading token, so "AMP (SUJU)" -> AMP.

USAGE
    python tools/build_quota_ledger.py <master.xlsx> [-o out.json] [--check ref.json]

--check compares against an existing ledger instead of writing: use it to prove
the generator reproduces a known-good file before trusting it on new input.
"""
import argparse
import json
import os
import re
import sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    sys.exit("needs openpyxl:  pip install openpyxl")

NAME_ROW, HS_ROW = 2, 3
COL_NO, COL_COMPANY, COL_LABEL = 1, 2, 3
TOTAL_HDR = "JUMLAH (MT)"


def clean(v):
    return str(v).strip() if v is not None else ""


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = clean(v).replace(",", "")
    if not s or s.upper() in ("TBA", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def build(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Status Submisson"]

    # HS-coded product columns: every column whose row-3 cell looks like an HS code
    cols = OrderedDict()          # col -> hs
    products = OrderedDict()      # hs  -> product name
    for c in range(1, ws.max_column + 1):
        hs = clean(ws.cell(row=HS_ROW, column=c).value)
        if not re.match(r"^\d{4}[.]?\d{2}[.]?\d{2}$|^\d{8,10}$", hs.replace(" ", "")):
            continue
        name = clean(ws.cell(row=NAME_ROW, column=c).value)
        if clean(name) == TOTAL_HDR:
            continue
        name = re.sub(r"\s+", " ", name).strip()
        cols[c] = hs
        products.setdefault(hs, name)

    companies = OrderedDict()
    cur = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=COL_NO).value
        b = clean(ws.cell(row=r, column=COL_COMPANY).value)
        if isinstance(a, (int, float)) and b:
            tok = re.split(r"[\s(]+", b)[0].upper()
            if re.match(r"^[A-Z]{2,6}$", tok):
                cur = tok
                companies.setdefault(cur, OrderedDict())
        if not cur:
            continue
        lab = clean(ws.cell(row=r, column=COL_LABEL).value)
        if re.match(r"^Obtained", lab, re.I) or re.match(r"^Revision", lab, re.I):
            field = "obtained"
        elif re.match(r"^Utili", lab, re.I) and "(MT)" in lab:
            field = "util"
        else:
            continue
        for c, hs in cols.items():
            v = num(ws.cell(row=r, column=c).value)
            if v == 0:
                continue
            ent = companies[cur].setdefault(hs, {"obtained": 0.0, "util": 0.0})
            ent[field] += v
    wb.close()

    # drop empties, normalise -0.0 and integral floats
    out = OrderedDict()
    for co, ent in companies.items():
        keep = OrderedDict()
        for hs, v in ent.items():
            o, u = round(v["obtained"], 3) + 0.0, round(v["util"], 3) + 0.0
            if o == 0 and u == 0:
                continue
            keep[hs] = {"obtained": int(o) if o == int(o) else o,
                        "util": int(u) if u == int(u) else u}
        if keep:
            out[co] = keep
    used = {hs for ent in out.values() for hs in ent}
    return {"products": OrderedDict((k, v) for k, v in products.items() if k in used),
            "companies": out}


def totals(led):
    o = sum(v["obtained"] for ent in led["companies"].values() for v in ent.values())
    u = sum(v["util"] for ent in led["companies"].values() for v in ent.values())
    return o, u


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("master")
    ap.add_argument("-o", "--out")
    ap.add_argument("--check", help="compare against this ledger instead of writing")
    a = ap.parse_args()

    led = build(a.master)
    o, u = totals(led)
    print(f"parsed {len(led['companies'])} companies, {len(led['products'])} products")
    print(f"  obtained {o:,.0f}   util {u:,.0f}   available {o - u:,.0f}")

    if a.check:
        ref = json.load(open(a.check, encoding="utf-8"))
        ro, ru = totals(ref)
        print(f"reference: obtained {ro:,.0f}   util {ru:,.0f}")
        diffs = []
        for co in sorted(set(led["companies"]) | set(ref["companies"])):
            g, e = led["companies"].get(co, {}), ref["companies"].get(co, {})
            for hs in sorted(set(g) | set(e)):
                gv, ev = g.get(hs, {"obtained": 0, "util": 0}), e.get(hs, {"obtained": 0, "util": 0})
                if gv["obtained"] != ev["obtained"] or gv["util"] != ev["util"]:
                    diffs.append(f"  {co:5} {hs:12} generated {gv} != reference {ev}")
        print(f"\nper-(company,HS) mismatches: {len(diffs)}")
        for d in diffs[:40]:
            print(d)
        return 0 if not diffs else 1

    led_out = OrderedDict()
    led_out["_meta"] = {
        "source": "master " + os.path.basename(a.master),
        "generated": os.environ.get("LEDGER_DATE", ""),
        "view": "effective (obtained incl. revisions)",
        "generator": "tools/build_quota_ledger.py",
    }
    led_out["products"] = led["products"]
    led_out["companies"] = led["companies"]
    dst = a.out or "iqdash/data/quotaLedger.json"
    with open(dst, "w", encoding="utf-8") as fh:
        json.dump(led_out, fh, indent=1, ensure_ascii=False)
        fh.write("\n")
    print(f"wrote {dst}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
